import openpyxl
from user import User
from post import Post

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]
product_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_no = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # Calculate number of products per supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        product_per_supplier[supplier_name] = 1

    # Calculate total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_cost = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_cost + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic products with inventory less than 1
    if inventory < 10:
        products_under_10_inv[int(product_no)] = int(inventory)

    # Add value for total inventory price
    inventory_price.value = inventory * price

print(product_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inventory_file.save("inventory_with_total_value.xlsx")

user = User("kockpaul@gmail.com", "Paul", "Password", "Developer")
user.get_user_info()
user.change_job_title("Devops Engineer")
user.get_user_info()

post = Post("Hello from the edge of the world", user.name)
post.get_post_info()

